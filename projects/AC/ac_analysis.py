""" CA Analysis Main """

#Note: to add new filter rules see readme.md file.

import sys
import os

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from dataclasses import dataclass
from typing import Callable, Any
from datetime import date
from functools import lru_cache


""" Columns (validate they match the Excel Sheet) """

CARDHOLDER    = "B"
MERCHANT_NAME = "C"
DOC_DESC      = "D"
TRANS_DATE    = "E"
AMMOUNT       = "F"
BT_MCC_CODE   = "O"
SAKNR_GL_ACCT = "AF"
CRITERIA_COL  = "AJ"  


""" AC utility functions and classes """

def get_letter_index(letter):
    """ expects a capital letter in the range (A - Z) """

    return ord(letter) - ord('A') + 1


def get_column_index(column):
    """ expects an excel column index a string of capital letters """

    base = get_letter_index('Z')
    index = 0
    exp = 0
    for col in column[::-1]:
        index += get_letter_index(col) * (base ** exp)
        exp += 1
    return index


def levenshtein_distance(a, b):

    len_a = len(a)
    len_b = len(b)

    @lru_cache(None)
    def min_dist(s1, s2):

        if s1 == len_a or s2 == len_b:
            return len_a - s1 + len_b - s2

        if a[s1] == b[s2]:
            return min_dist(s1 + 1, s2 + 1)

        return 1 + min(
            min_dist(s1, s2 + 1),      # cost of insert
            min_dist(s1 + 1, s2),      # cost of delete
            min_dist(s1 + 1, s2 + 1),  # cost of replace
        )

    return min_dist(0, 0)


def get_text_similarity(a: str, b: str) -> float:
    """ Calculates the similarity in percentage of two strings """

    a1 = a if a else ""
    b1 = b if b else ""
    max_len = max(1, len(a1), len(b1))
    dist = levenshtein_distance(a1, b1)
    return 100.0 * ((max_len - dist) / max_len)


@dataclass()
class ACRow:
    """ Wrapper class for an Excel rows """

    row: Any
    index: int = 0

    def __getitem__(self, column):
        index = get_column_index(column)
        return self.row[index - 1].value

    def cell(self, column):
        index = get_column_index(column)
        return self.row[index - 1]

    def highlight(self, color, column, criteria):

        for cell in self.row:
            cell.fill = color

        current_criteria = self[column]
        self.cell(column).value = f"{current_criteria} AND {criteria}" if current_criteria else criteria


@dataclass()
class ACContext:
    fps_cardholders: set[str]
    rcm_cardholders: set[str]
    dsc_list: set[str]
    highlight_colors: dict
    spend_max: float
    gl_codes: list[str]
    result_column: str

    def in_dsc_list(self, merchant) -> bool:
        return (merchant if merchant else "").upper() in self.dsc_list


@dataclass
class Transaction:
    row: ACRow

    @property
    def index(self)-> int: return self.row.index
    @property
    def cardholder(self) -> str: return self.row[CARDHOLDER]
    @property
    def merchant(self) -> str: return self.row[MERCHANT_NAME]
    @property
    def desc(self) -> str: return self.row[DOC_DESC]
    @property
    def date(self) -> date: return self.row[TRANS_DATE]
    @property
    def total(self) -> float: return self.row[AMMOUNT]


class Split:
    transactions: list[Transaction]
    index: int = 0

    def __init__(self, trans: Transaction) -> None:
        self.transactions = [trans]
        self.index = trans.index

    def try_add(self, trans: Transaction, max_days: int, similarity: float) -> bool:
        # compare with the first transaction
        master = self.transactions[0]

        # compare dates (ignore blanks)
        delta_day = (trans.date - master.date).days if trans.date and master.date else 0
        if delta_day > max_days:
            return False

        # compare merchants
        if get_text_similarity(master.merchant, trans.merchant) < similarity:
            return False

        self.transactions.append(trans)
        
        return True

    def equal_or_more(self, limit: float):
        return sum(map(lambda t: t.total, self.transactions)) >= limit

    def is_multiple(self) -> bool:
        return len(self.transactions) > 1


class CardHolder:
    name: str
    splits: list[Split]

    def __init__(self, name) -> None:
        self.name = name
        self.splits = []

    def add(self, trans: Transaction) -> bool:
        
        for split in self.splits:
            # try to add split with 4 days gap and 60% similarity on merchant name
            if split.try_add(trans, 4, 60.0):
                return False
        
        self.splits.append(Split(trans))

        return True

    def get_valid_splits(self, limit: float) -> list[Split]:
        return list(filter(lambda s: s.is_multiple() and s.equal_or_more(limit), self.splits))
        

def update_splits(cardholders: dict, row: ACRow):

    ch_name = row[CARDHOLDER]
    transaction = Transaction(row=row)

    if ch_name not in cardholders:
        cardholder = CardHolder(ch_name)
        cardholders[ch_name] = cardholder
    else:
        cardholder = cardholders[ch_name]

    cardholder.add(transaction)
    

def get_valid_splits(cardholders: dict, min_value: float) -> list[Split]:
    result = list()

    for key in cardholders:
        result.extend(cardholders[key].get_valid_splits(min_value))

    return sorted(result, key=lambda s: s.index)


def get_selected_transactions(splits: list[Split]) -> list[Transaction]:
    result = list()

    for s in splits:
        result.extend(s.transactions)

    return result


""" CA Analysis Engine """

class Filter:
    name: str
    rule: Callable
    color: str
    rows: list

    def __init__(self, name: str, rule: Callable, color: str) -> None:
        self.name = name
        self.rule = rule
        self.color = color
        self.rows = list()


###### filter rules - start ######

def mcc_9399(row: ACRow, context: ACContext) -> bool:
    return row[BT_MCC_CODE] == '9399'

def restricted_travel_expense(row: ACRow, context: ACContext) -> bool:

    gls = ["51300", "51306", "51312", "50928", "51318", "51302", "51308", "51314", "50951", "51320", 
           "51304", "51310", "51315", "50944", "51322", "50953", "51326", "51400", "51406", "51412", 
           "51418", "51424", "51402", "51408", "51414", "51420", "51426", "51404", "51410", "51416", 
           "51422", "51428", "51430", "51328", "50916", "50907", "10306", "50957", "50960", "50962", 
           "10312", "10330", "54994", "55995"]

    mcc = ["3009", "3351", "3357", "3366", "3389", "3393", "3395", "3405", "3501", "3502", "3503", 
           "3504", "3508", "3509", "3512", "3513", "3520", "3526", "3530", "3562", "3573", "3581", 
           "3590", "3615", "3637", "3640", "3649", "3665", "3687", "3690", "3692", "3703", "3715", 
           "3740", "3750", "3751", "3778", "4011", "4111", "4121", "4131", "4582", "4722", "4784", 
           "4789", "7011", "7033", "7512", "7513", "7519"]

    return row[SAKNR_GL_ACCT] in gls or row[BT_MCC_CODE] in mcc 

def mcc_7932(row: ACRow, context: ACContext) -> bool:
    return row[BT_MCC_CODE] == '7932'

def ta_limit_exceeded(row: ACRow, context: ACContext) -> bool:
    return row[AMMOUNT] >= context.spend_max

def convenience_cheques(row: ACRow, context: ACContext) -> bool:
    return (row[MERCHANT_NAME] == "CHEQUE" and row[AMMOUNT] is None) or row[BT_MCC_CODE] == "9991"


def gl_financial_coding_error(row: ACRow, context: ACContext) -> bool:
    return row[SAKNR_GL_ACCT] in ['52334', '52136'] and context.in_dsc_list(row[MERCHANT_NAME])

def gl_conference_fees(row: ACRow, context: ACContext) -> bool:
    return row[SAKNR_GL_ACCT] in ['52334', '52136'] and not context.in_dsc_list(row[MERCHANT_NAME])

def gl_hospitality(row: ACRow, context: ACContext) -> bool:

    gls = ["52204", "52206", "52207", "54211"]
    mcc = ["5411", "5422", "5441", "5462", "5499", "5811", "5812", "5813", "5814", "5921", 
           "5992", "5993", "7922", "7929", "7932", "7941", "7991", "7998", "7999", "8398"]
    
    return row[SAKNR_GL_ACCT] in gls or row[BT_MCC_CODE] in mcc

def gl_restricted_memberships(row: ACRow, context: ACContext) -> bool:
    return row[SAKNR_GL_ACCT] in ['58208', '52336', '52337'] or row[BT_MCC_CODE] in ["7997", "8641", "8699"]

def gl_restricted_motor_vehicles(row: ACRow, context: ACContext) -> bool:

    gls = ["53332", "53334", "54285", "54287" "54000", "54002", "54004"]
    mcc = ["5013", "5172", "5511", "5532", "5533", "5541", "5542", "5599", "7523", 
           "7531", "7534", "7535", "7538", "7542", "7549", "7692", "7699", "8675"]

    return row[SAKNR_GL_ACCT] in gls or row[BT_MCC_CODE] in mcc

def rcm_cardholders_vlookup(row: ACRow, context: ACContext) -> bool:
    return row[CARDHOLDER] in context.rcm_cardholders

def restricted_home_equipment_purchases(row: ACRow, context: ACContext) -> bool:
    
    gls = ["54260", "54206", "54208"]
    mcc = ["5021", "5712", "5719", "7641"]

    return (row[SAKNR_GL_ACCT] in gls or row[BT_MCC_CODE] in mcc) and row[AMMOUNT] >= 800.00 and row[CARDHOLDER] not in context.fps_cardholders

def restricted_it_purchases(row: ACRow, context: ACContext) -> bool:    

    gls = ["53150", "52385", "52386", "52387", "53140", "53142", "53362", "54312", "54295", "54040", "54042", "54050"]
    mcc = ["4816", "5045", "5732", "5734", "5815", "5816", "5817", "5818", "7379", "7622"]

    return (row[SAKNR_GL_ACCT] in gls or row[BT_MCC_CODE] in mcc) and row[AMMOUNT] >= 800.00 and row[CARDHOLDER] not in context.fps_cardholders

def books_papers_subs(row: ACRow, context: ACContext) -> bool: 
    return row[BT_MCC_CODE] in ["5192", "5942", "5994"]

def fines_and_penalties(row: ACRow, context: ACContext) -> bool: 
    return row[BT_MCC_CODE] in ["9222", "9311"]

def personal_purchases(row: ACRow, context: ACContext) -> bool: 
    
    mcc = ["5094", "5733", "5932", "5940", "5944", "5945", "5947", "5948", "5950", "5970", 
           "5971", "5972", "5977", "5997", "7230", "7298", "7333", "7841", "7911"]

    return row[BT_MCC_CODE] in mcc


###### filter rules - end ######

def get_filters():
    filters = [
        Filter(name="Restricted Home Equipment Purchases", rule=restricted_home_equipment_purchases, color="Olive"),
        Filter(name="Restricted IS/OGD transactions", rule=mcc_9399, color="Cornsilk"), 
        Filter(name="Restricted Travel Status Related Expenses", rule=restricted_travel_expense, color="NavajoWhite"),
        Filter(name="Event/Hospitality", rule=mcc_7932, color="Salmon"),
        Filter(name="TA Limit exceeded", rule=ta_limit_exceeded, color="DarkSalmon"),
        Filter(name="Monitoring of Convenience Cheques", rule=convenience_cheques, color="RosyBrown"),
        Filter(name="Financial Coding Error - Designated Science Conferences", rule=gl_financial_coding_error, color="Pink"),
        Filter(name="Conference fees", rule=gl_conference_fees, color="HotPink"),
        Filter(name="Hospitality", rule=gl_hospitality, color="Gold"),
        Filter(name="Restricted Memberships", rule=gl_restricted_memberships, color="Violet"),
        Filter(name="Restricted Motor Vehicles Expenses", rule=gl_restricted_motor_vehicles, color="DarkSeaGreen"),
        Filter(name="SoD - TA and s.34", rule=rcm_cardholders_vlookup, color="Aqua"),
        Filter(name="Restricted - IT Purchases", rule=restricted_it_purchases, color="Aquamarine"),
        Filter(name="Books, Newspapers, and Subscriptions", rule=books_papers_subs, color="Bisque"),
        Filter(name="Fines and Penalties", rule=fines_and_penalties, color="Ivory"),
        Filter(name="Personal Purchases", rule=personal_purchases, color="Moccasin")
    ]
    return filters


class Analizer:

    filters: list[Filter]
    context: ACContext
    cardholders: dict

    def __init__(self, context: ACContext) -> None:
        self.filters = get_filters()
        self.context = context
        self.cardholders = {}

    def analize(self, row: ACRow):
        
        # add row to all filters that match
        for f in self.filters:
            if f.rule(row, self.context):
                f.rows.append(row)    

        # update splits
        if row[CARDHOLDER] not in self.context.fps_cardholders and row[SAKNR_GL_ACCT] not in self.context.gl_codes:
            update_splits(self.cardholders, row)

    def highlight_rows(self):

        # highlight rules
        self.highlight_filters()
        
        # highlight splits
        self.highlight_splits()

    def highlight_filters(self):
        for f in self.filters:
            color = self.context.highlight_colors[f.color]
            for r in f.rows:
                r.highlight(color, self.context.result_column, f.name)

    def highlight_splits(self):
        
        color = self.context.highlight_colors["OrangeRed"]

        splits = get_valid_splits(self.cardholders, self.context.spend_max)
        trans = get_selected_transactions(splits)

        for t in trans:
            t.row.highlight(color, self.context.result_column, "Potential Splits")

    def __str__(self) -> str:
        return "\n".join(map(lambda f: f"{f.name}: {len(f.rows)}", self.filters))


def _get_output_name(file_name: str):
    split = os.path.splitext(file_name)
    return f"{split[0]}_results{split[1]}"


def _get_pattern(rgb: str) -> PatternFill:
    return PatternFill(start_color=rgb, end_color=rgb, fill_type="solid")


def _get_highlight_colors() -> dict:
    return {
        "Cornsilk": _get_pattern("FFF8DC"),
        "NavajoWhite": _get_pattern("FFDEAD"), 
        "Salmon": _get_pattern("FA8072"), 
        "DarkSalmon": _get_pattern("E9967A"),
        "RosyBrown": _get_pattern("BC8F8F"),
        "Pink": _get_pattern("FFC0CB"),
        "HotPink": _get_pattern("FF69B4"),
        "Coral": _get_pattern("FF7F50"),
        "OrangeRed": _get_pattern("FF4500"),
        "DarkSeaGreen": _get_pattern("8FBC8F"),
        "Aqua": _get_pattern("00FFFF"),
        "Violet": _get_pattern("EE82EE"),
        "Olive": _get_pattern("808000"),
        "Gold": _get_pattern("FFD700"),
        "Aquamarine": _get_pattern("7FFFD4"),
        "Bisque": _get_pattern("FFE4C4"),
        "Ivory": _get_pattern("FFFFF0"),
        "Moccasin": _get_pattern("FFE4B5")
    }


def _load_set_column(sheet, from_row, col) -> set[str]:
    values = set()
    for row in sheet.iter_rows(min_row = from_row, max_col = col, max_row = sheet.max_row):
        if row[col - 1].value:
            values.add(row[col - 1].value.strip().upper())
    return values


def _load_fps_chs(sheet) -> set[str]:
    return _load_set_column(sheet, 2, 2)


def _load_rcm_chs(sheet) -> set[str]:
    return _load_set_column(sheet, 2, 3)


def _load_dsc_list(sheet) -> set[str]:
    return _load_set_column(sheet, 2, 3)


def _get_context(workbook) -> ACContext:

    fps_chs = _load_fps_chs(workbook["FPS CHs"])
    rcm_chs = _load_rcm_chs(workbook["RCM CHs"])
    dsc_list = _load_dsc_list(workbook["DSC"])
    colors = _get_highlight_colors()
    gl_codes = ['52017', '52008', '52148', '54374', '52136', '52334']

    return ACContext(fps_cardholders=fps_chs, rcm_cardholders=rcm_chs, dsc_list=dsc_list, 
        highlight_colors=colors, result_column=CRITERIA_COL, spend_max=10000.0, gl_codes=gl_codes)


def _validate_sheets(names: list[str]) -> bool:
    return "Master" in names and "FPS CHs" in names and "RCM CHs" in names and "DSC" in names


def run_analysis(argvs):

    if len(argvs) != 2:
        _, tail = os.path.split(argvs[0])
        print(f'Usage:\n\t{tail} <excel_file_name>')
        return 0

    input_file = argvs[1]

    print(f"Loading file...")

    workbook = load_workbook(input_file, read_only=False, data_only=True)

    print(f"Validating sheets...")

    if not _validate_sheets(workbook.sheetnames):
        print("One of these sheets was not found: 'Master', 'FPS CHs', 'RCM CHs', 'DSC'.\nPlease rename the sheets to be able to run the analysis.")
        return 0

    master_sheet = workbook["Master"]
    
    print(f"Analysing...")

    context = _get_context(workbook)
    last_column = master_sheet.max_column + 1
    criteria_column = get_column_index(CRITERIA_COL)
    
    master_sheet.cell(row=1, column=criteria_column).value = "CRITERIA"

    analyzer = Analizer(context)

    # iterates the master sheet
    row_index = 2
    for excel_row in master_sheet.iter_rows(min_row=row_index, max_col=last_column, max_row=master_sheet.max_row):

        row = ACRow(row=excel_row, index=row_index)

        if row[CARDHOLDER] is not None:
            analyzer.analize(row)

        row_index += 1

    print(f"Highlighting rows...")

    # highlight the rows
    analyzer.highlight_rows()

    #print(analyzer)
    
    # save the file
    output_file = _get_output_name(input_file)
    workbook.save(output_file)

    print(f"Process completed...\nOutput file: {output_file}\n")

    return 0


if __name__ == '__main__':
    sys.exit(run_analysis(sys.argv))
