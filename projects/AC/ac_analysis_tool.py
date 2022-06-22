""" CA Analysis Main """

#Note: to add new filter rules see readme.md file.

import sys
import os

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from dataclasses import dataclass
from typing import Callable, Any
from datetime import date
from functools import lru_cache


""" Columns (validate they match the Excel Sheet) """

CARDHOLDER    = "B"
MERCHANT_NAME = "C"
DOC_DESC      = "D"
TRANS_DATE    = "E"
AMMOUNT       = "F"
BT_MCC_CODE   = "O"
SAKNR_GL_ACCT = "AF"
CRITERIA_COL  = "AJ"  


""" AC utility functions and classes """

def get_letter_index(letter):
    """ expects a capital letter in the range (A - Z) """

    return ord(letter) - ord('A') + 1


def get_column_index(column):
    """ expects an excel column index a string of capital letters """

    base = get_letter_index('Z')
    index = 0
    exp = 0
    for col in column[::-1]:
        index += get_letter_index(col) * (base ** exp)
        exp += 1
    return index


def levenshtein_distance(a, b):

    len_a = len(a)
    len_b = len(b)

    @lru_cache(None)
    def min_dist(s1, s2):

        if s1 == len_a or s2 == len_b:
            return len_a - s1 + len_b - s2

        if a[s1] == b[s2]:
            return min_dist(s1 + 1, s2 + 1)

        return 1 + min(
            min_dist(s1, s2 + 1),      # cost of insert
            min_dist(s1 + 1, s2),      # cost of delete
            min_dist(s1 + 1, s2 + 1),  # cost of replace
        )

    return min_dist(0, 0)


def get_text_similarity(a: str, b: str) -> float:
    """ Calculates the similarity in percentage of two strings """

    a1 = a if a else ""
    b1 = b if b else ""
    max_len = max(1, len(a1), len(b1))
    dist = levenshtein_distance(a1, b1)
    return 100.0 * ((max_len - dist) / max_len)


@dataclass()
class ACRow:
    """ Wrapper class for an Excel rows """

    row: Any
    index: int = 0

    def __getitem__(self, column):
        index = get_column_index(column)
        return self.row[index - 1].value

    def cell(self, column):
        index = get_column_index(column)
        return self.row[index - 1]

    def highlight(self, color, column, criteria):

        for cell in self.row:
            cell.fill = color

        current_criteria = self[column]
        self.cell(column).value = f"{current_criteria} AND {criteria}" if current_criteria else criteria


@dataclass()
class ACContext:
    fps_cardholders: set[str]
    rcm_cardholders: set[str]
    dsc_list: set[str]
    highlight_colors: dict
    spend_max: float
    gl_codes: list[str]
    result_column: str

    def in_dsc_list(self, merchant) -> bool:
        return (merchant if merchant else "").upper() in self.dsc_list


@dataclass
class Transaction:
    row: ACRow

    @property
    def index(self)-> int: return self.row.index
    @property
    def cardholder(self) -> str: return self.row[CARDHOLDER]
    @property
    def merchant(self) -> str: return self.row[MERCHANT_NAME]
    @property
    def desc(self) -> str: return self.row[DOC_DESC]
    @property
    def date(self) -> date: return self.row[TRANS_DATE]
    @property
    def total(self) -> float: return self.row[AMMOUNT]


class Split:
    transactions: list[Transaction]
    index: int = 0

    def __init__(self, trans: Transaction) -> None:
        self.transactions = [trans]
        self.index = trans.index

    def try_add(self, trans: Transaction, max_days: int, similarity: float) -> bool:
        # compare with the first transaction
        master = self.transactions[0]

        # compare dates (ignore blanks)
        delta_day = (trans.date - master.date).days if trans.date and master.date else 0
        if delta_day > max_days:
            return False

        # compare merchants
        if get_text_similarity(master.merchant, trans.merchant) < similarity:
            return False

        self.transactions.append(trans)
        
        return True

    def equal_or_more(self, limit: float):
        return sum(map(lambda t: t.total, self.transactions)) >= limit

    def is_multiple(self) -> bool:
        return len(self.transactions) > 1


class CardHolder:
    name: str
    splits: list[Split]

    def __init__(self, name) -> None:
        self.name = name
        self.splits = []

    def add(self, trans: Transaction) -> bool:
        
        for split in self.splits:
            # try to add split with 4 days gap and 60% similarity on merchant name
            if split.try_add(trans, 4, 60.0):
                return False
        
        self.splits.append(Split(trans))

        return True

    def get_valid_splits(self, limit: float) -> list[Split]:
        return list(filter(lambda s: s.is_multiple() and s.equal_or_more(limit), self.splits))
        

def update_splits(cardholders: dict, row: ACRow):

    ch_name = row[CARDHOLDER]
    transaction = Transaction(row=row)

    if ch_name not in cardholders:
        cardholder = CardHolder(ch_name)
        cardholders[ch_name] = cardholder
    else:
        cardholder = cardholders[ch_name]

    cardholder.add(transaction)
    

def get_valid_splits(cardholders: dict, min_value: float) -> list[Split]:
    result = list()

    for key in cardholders:
        result.extend(cardholders[key].get_valid_splits(min_value))

    return sorted(result, key=lambda s: s.index)


def get_selected_transactions(splits: list[Split]) -> list[Transaction]:
    result = list()

    for s in splits:
        result.extend(s.transactions)

    return result


""" CA Analysis Engine """

class Filter:
    name: str
    rule: Callable
    color: str
    rows: list

    def __init__(self, name: str, rule: Callable, color: str) -> None:
        self.name = name
        self.rule = rule
        self.color = color
        self.rows = list()


###### filter rules - start ######

def mcc_9399(row: ACRow, context: ACContext) -> bool:
    """ Restricted IS/OGD transactions => column O = 9399 """
    return row[BT_MCC_CODE] == '9399'

def restricted_travel_expense(row: ACRow, context: ACContext) -> bool:
    """ Restricted - Travel Status Related Expenses """
    return row[BT_MCC_CODE] in ['7011', '4111', '7513'] or row[SAKNR_GL_ACCT] in ["54260", "54206", "54208"]

def mcc_7932(row: ACRow, context: ACContext) -> bool:
    """ Event/Hospitality => column O = 7932 """
    return row[BT_MCC_CODE] == '7932'

def ta_limit_exceeded(row: ACRow, context: ACContext) -> bool:
    """ TA Limit exceeded => column F > MAX """
    return row[AMMOUNT] >= context.spend_max

def empty_cheques(row: ACRow, context: ACContext) -> bool:
    """ Monitoring of Convenience Cheques => column C = CHEQUE and F is empty """
    return row[MERCHANT_NAME] == "CHEQUE" and row[AMMOUNT] is None

def gl_financial_coding_error(row: ACRow, context: ACContext) -> bool:
    """ Financial Coding Error - Designated Science Conferences => column AF = 52334 or 52136 and B in DSC list"""
    return row[SAKNR_GL_ACCT] in ['52334', '52136'] and context.in_dsc_list(row[MERCHANT_NAME])

def gl_conference_fees(row: ACRow, context: ACContext) -> bool:
    """ Conference fees => column AF = 52334 or 52136 and C not in DSC list """
    return row[SAKNR_GL_ACCT] in ['52334', '52136'] and not context.in_dsc_list(row[MERCHANT_NAME])

def gl_hospitality(row: ACRow, context: ACContext) -> bool:
    """ Hospitality => column AF = 52204, 52206, 52207, 54211 """
    return row[SAKNR_GL_ACCT] in ['52204', '52206', '52207', '54211']

def gl_restricted_memberships(row: ACRow, context: ACContext) -> bool:
    """ Restricted Memberships => column AF = 58208, 52336, 52337 """
    return row[SAKNR_GL_ACCT] in ['58208', '52336', '52337']

def gl_restricted_motor_vehicles(row: ACRow, context: ACContext) -> bool:
    """ Restricted Motor Vehicles Expenses => column AF = 53332, 53334, 54285, 54287 54000, 54002, 54004  """ 
    return row[SAKNR_GL_ACCT] in ["53332", "53334", "54285", "54287" "54000", "54002", "54004"]

def gl_restricted_home_equipment_purchases(row: ACRow, context: ACContext) -> bool:
    """ Restricted Home Equipment Purchases => AF = 54260 and F >= $800 and CH not in FPS-CHs list """
    return row[SAKNR_GL_ACCT] == "54260" and row[AMMOUNT] >= 800.00 and row[CARDHOLDER] not in context.fps_cardholders

def rcm_cardholders_vlookup(row: ACRow, context: ACContext) -> bool:
    """ SoD - TA and s.34 => CH in RCM-CHs list """
    return row[CARDHOLDER] in context.rcm_cardholders

###### filter rules - end ######


def get_filters():
    filters = [
        Filter(name="Restricted IS/OGD transactions", rule=mcc_9399, color="Cornsilk"), 
        Filter(name="Restricted Travel Status Related Expenses", rule=restricted_travel_expense, color="NavajoWhite"),
        Filter(name="Event/Hospitality", rule=mcc_7932, color="Salmon"),
        Filter(name="TA Limit exceeded", rule=ta_limit_exceeded, color="DarkSalmon"),
        Filter(name="Monitoring of Convenience Cheques", rule=empty_cheques, color="RosyBrown"),
        Filter(name="Financial Coding Error - Designated Science Conferences", rule=gl_financial_coding_error, color="Pink"),
        Filter(name="Conference fees", rule=gl_conference_fees, color="HotPink"),
        Filter(name="Hospitality", rule=gl_hospitality, color="Gold"),
        Filter(name="Restricted Memberships", rule=gl_restricted_memberships, color="Violet"),
        Filter(name="Restricted Home Equipment Purchases", rule=gl_restricted_home_equipment_purchases, color="Olive"),
        Filter(name="Restricted Motor Vehicles Expenses", rule=gl_restricted_motor_vehicles, color="DarkSeaGreen"),
        Filter(name="SoD - TA and s.34", rule=rcm_cardholders_vlookup, color="Aqua")
    ]
    return filters


class Analizer:

    filters: list[Filter]
    context: ACContext
    cardholders: dict

    def __init__(self, context: ACContext) -> None:
        self.filters = get_filters()
        self.context = context
        self.cardholders = {}

    def analize(self, row: ACRow):
        
        # add row to all filters that match
        for f in self.filters:
            if f.rule(row, self.context):
                f.rows.append(row)    

        # update splits
        if row[CARDHOLDER] not in self.context.fps_cardholders and row[SAKNR_GL_ACCT] not in self.context.gl_codes:
            update_splits(self.cardholders, row)

    def highlight_rows(self):

        # highlight rules
        self.highlight_filters()
        
        # highlight splits
        self.highlight_splits()

    def highlight_filters(self):
        for f in self.filters:
            color = self.context.highlight_colors[f.color]
            for r in f.rows:
                r.highlight(color, self.context.result_column, f.name)

    def highlight_splits(self):
        
        color = self.context.highlight_colors["OrangeRed"]

        splits = get_valid_splits(self.cardholders, self.context.spend_max)
        trans = get_selected_transactions(splits)

        for t in trans:
            t.row.highlight(color, self.context.result_column, "Potential Splits")

    def __str__(self) -> str:
        return "\n".join(map(lambda f: f"{f.name}: {len(f.rows)}", self.filters))


def _get_output_name(file_name: str):
    split = os.path.splitext(file_name)
    return f"{split[0]}_results{split[1]}"


def _get_pattern(rgb: str) -> PatternFill:
    return PatternFill(start_color=rgb, end_color=rgb, fill_type="solid")


def _get_highlight_colors() -> dict:
    return {
        "Cornsilk": _get_pattern("FFF8DC"),
        "NavajoWhite": _get_pattern("FFDEAD"), 
        "Salmon": _get_pattern("FA8072"), 
        "DarkSalmon": _get_pattern("E9967A"),
        "RosyBrown": _get_pattern("BC8F8F"),
        "Pink": _get_pattern("FFC0CB"),
        "HotPink": _get_pattern("FF69B4"),
        "Coral": _get_pattern("FF7F50"),
        "OrangeRed": _get_pattern("FF4500"),
        "DarkSeaGreen": _get_pattern("8FBC8F"),
        "Aqua": _get_pattern("00FFFF"),
        "Violet": _get_pattern("EE82EE"),
        "Olive": _get_pattern("808000"),
        "Gold": _get_pattern("FFD700")
    }


def _load_set_column(sheet, from_row, col) -> set[str]:
    values = set()
    for row in sheet.iter_rows(min_row = from_row, max_col = col, max_row = sheet.max_row):
        if row[col - 1].value:
            values.add(row[col - 1].value.strip().upper())
    return values


def _load_fps_chs(sheet) -> set[str]:
    return _load_set_column(sheet, 2, 2)


def _load_rcm_chs(sheet) -> set[str]:
    return _load_set_column(sheet, 2, 3)


def _load_dsc_list(sheet) -> set[str]:
    return _load_set_column(sheet, 2, 3)


def _get_context(workbook) -> ACContext:

    fps_chs = _load_fps_chs(workbook["FPS CHs"])
    rcm_chs = _load_rcm_chs(workbook["RCM CHs"])
    dsc_list = _load_dsc_list(workbook["DSC"])
    colors = _get_highlight_colors()
    gl_codes = ['52017', '52008', '52148', '54374', '52136', '52334']

    return ACContext(fps_cardholders=fps_chs, rcm_cardholders=rcm_chs, dsc_list=dsc_list, 
        highlight_colors=colors, result_column=CRITERIA_COL, spend_max=10000.0, gl_codes=gl_codes)


def _validate_sheets(names: list[str]) -> bool:
    return "Master" in names and "FPS CHs" in names and "RCM CHs" in names and "DSC" in names


def run_analysis(argvs):

    if len(argvs) != 2:
        _, tail = os.path.split(argvs[0])
        print(f'Usage:\n\t{tail} <excel_file_name>')
        return 0

    input_file = argvs[1]

    print(f"Loading file...")

    workbook = load_workbook(input_file, read_only=False, data_only=True)

    print(f"Validating sheets...")

    if not _validate_sheets(workbook.sheetnames):
        print("One of these sheets was not found: 'Master', 'FPS CHs', 'RCM CHs', 'DSC'.\nPlease rename the sheets to be able to run the analysis.")
        return 0

    master_sheet = workbook["Master"]
    
    print(f"Analysing...")

    context = _get_context(workbook)
    last_column = master_sheet.max_column + 1
    criteria_column = get_column_index(CRITERIA_COL)
    
    master_sheet.cell(row=1, column=criteria_column).value = "CRITERIA"

    analyzer = Analizer(context)

    # iterates the master sheet
    row_index = 2
    for excel_row in master_sheet.iter_rows(min_row=row_index, max_col=last_column, max_row=master_sheet.max_row):

        row = ACRow(row=excel_row, index=row_index)

        if row[CARDHOLDER] is not None:
            analyzer.analize(row)

        row_index += 1

    print(f"Highlighting rows...")

    # highlight the rows
    analyzer.highlight_rows()

    #print(analyzer)
    
    # save the file
    output_file = _get_output_name(input_file)
    workbook.save(output_file)

    print(f"Process completed...\nOutput file: {output_file}\n")

    return 0


if __name__ == '__main__':
    sys.exit(run_analysis(sys.argv))
