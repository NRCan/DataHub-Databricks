# requires python 3.8
# requires openpyxl package

from shutil import copyfile
from openpyxl import load_workbook
from openpyxl.styles import PatternFill 

def valid_number(v):
  return isinstance(v, float) or isinstance(v, int)

peach = "00FFCC99"
hilight_pattern = PatternFill(start_color=peach, end_color=peach, fill_type = "solid")

root_path = "/dbfs/FileStore/shared_uploads/hipolito.perez-garcia@nrcan-rncan.gc.ca/"
input_file = root_path + "goa_expenses.xlsx"
temp_file = "/tmp/temp.xlsx"
output_file = root_path + "goa_expenses_out.xlsx"

workbook = load_workbook(input_file, read_only=False, data_only=True)
sheet = workbook["goa_expenses"]

check_col_index = 6
update_col_index = 9
top_column = 11

for row in sheet.iter_rows(min_row=2, max_col=top_column, max_row = sheet.max_row):
  cell_value = row[check_col_index].value
  if valid_number(cell_value) and (cell_value > 1000.0):
    row[update_col_index].value = "Yes"
    #hilight row
    for cell in row: 
      cell.fill = hilight_pattern
  else:
    row[update_col_index].value = "No"

# save to a temp file where storage allows random writes
workbook.save(temp_file)

# copy to local working folder
copyfile(temp_file, output_file)