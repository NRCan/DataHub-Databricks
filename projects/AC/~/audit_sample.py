import sys
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill 

def is_column_letter(letter: str) -> bool:
    return (letter >= "A") and (letter <= "Z")

def get_column_index(letter: str) -> int:
    return ord(letter.upper()[0]) - ord('A')

def expression_column_mapper(c: str): 
    if is_column_letter(c): 
        return "row[{}].value".format(get_column_index(c))
    else:
        return c

def scan_expression(text: str) -> list[str]:
    output = []
    string = []
    for c in text:
        if string:
            string.append(c)
            if c == "'":
                output.append("".join(string))
                string = []
        else:
            if c == "'":
                string = [c]
            else:
                output.append(c)
    return output

def build_test_expression(template: str) -> str:
    return "".join(map(expression_column_mapper, scan_expression(template)))   

def get_output_name(input_name: str):
    split = os.path.splitext(input_name)
    return split[0] + "_results" + split[1]

def get_highlight_pattern():
    peach = "00FFCC99"
    return PatternFill(start_color=peach, end_color=peach, fill_type="solid")

def valid_condition(expression):
    try:
        compile(expression, "<string>", "eval")
        return True
    except:
        return False

def test_row(expression, row): 
    try:
        return eval(expression, { "row": row })    
    except:
        return None

def main(argvs) -> int:
    #validate inputs
    if len(argvs) != 4:
        head, tail = os.path.split(argvs[0])
        print('Usage:\n\t{} <excel_file_name> <last_data_column> "<filter_condition>"'.format(tail))
        print('E.g.:\n\t{} sample.xlsx I "G > 1000"'.format(tail))
        return 0

    if not valid_condition(argvs[3]): 
        print('ERROR: "{}" is not a valid Python expression!'.format(argvs[3]))
        return 0   

    hilight_pattern = get_highlight_pattern()

    input_file = argvs[1]
    output_file = get_output_name(input_file)
    update_col_index = get_column_index(argvs[2]) + 1
    test_expression = build_test_expression(argvs[3])

    workbook = load_workbook(input_file, read_only=False, data_only=True)
    sheet = workbook.worksheets[0]

    sheet.cell(row = 1, column = update_col_index + 1).value = "Results"

    for row in sheet.iter_rows(min_row = 2, max_col = update_col_index + 1, max_row = sheet.max_row):
        test_result = test_row(test_expression, row)
        if test_result == True:
            row[update_col_index].value = "Yes"
            #hilight row
            for cell in row: 
                cell.fill = hilight_pattern
        else:
            row[update_col_index].value = "No" if test_result != None else "INVALID"

    # save output file
    workbook.save(output_file)
    print(output_file)

    return 0

if __name__ == '__main__':
    sys.exit(main(sys.argv))
    #sys.exit(main([sys.argv[0], "goa_expenses_test.xlsx", "I", "G >= 200 and G <= 500"]))    

# to build the EXE use: > pyinstaller -F .\audit_sample.py        