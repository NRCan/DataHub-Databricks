""" CA Processor Rules """

from dataclasses import dataclass
from pyclbr import Function
from typing import Any, Callable
from openpyxl.styles import PatternFill
from ac_processor_utils import ACRow, ACContext

def _make_pattern(color: str) -> PatternFill:
    return PatternFill(start_color=color, end_color=color, fill_type="solid")

###### rules - start ######
def mcc_9399(row: ACRow, context: object) -> object:
    """ Restricted IS/OGD transactions => column N = 9399 """
    return _make_pattern("00FFCC99") if row["N"] == '9399' else None

def mcc_7011(row: ACRow, context: object) -> object:
    """ Restricted Travel Status Related Expenses => column N = 7011 """
    return _make_pattern("00FFCC90") if row["N"] == '7011' else None

def mcc_4111(row: ACRow, context: object) -> object:
    """ Restricted Travel Status Related Expenses => column N = 4111 """
    return _make_pattern("00FFCC89") if row["N"] == '4111' else None

def mcc_7513(row: ACRow, context: object) -> object:
    """ Restricted Travel Status Related Expenses => column N = 7513 """
    return _make_pattern("00FFCC80") if row["N"] == '7513' else None

def mcc_7932(row: ACRow, context: object) -> object:
    """ Event/Hospitality => column N = 7932 """
    return _make_pattern("00FFCC79") if row["N"] == '7932' else None

def ta_limit_exceeded(row: ACRow, context: object) -> object:
    """ TA Limit exceeded => column E > 5000 """
    return _make_pattern("00FFCC70") if row["E"] > 5000.00 else None

def empty_cheques(row: ACRow, context: object) -> object:
    """ Monitoring of Convenience Cheques => column B = CHEQUE and E is empty """
    return _make_pattern("00FFCC69") if row["B"] == "CHEQUE" else None

def gl_financial_coding_error(row: ACRow, context: object) -> object:
    """ Financial Coding Error - Designated Science Conferences => column AE = 52334 or 52136 """
    return _make_pattern("00FFCC60") if row["AE"] in ['52334', '52136'] else None

def gl_conference_fees(row: ACRow, context: object) -> object:
    """ Conference fees => column AE = 52334 or 52136 (review) """
    return _make_pattern("00FFCC59") if row["AE"] in ['52334', '52136'] else None

def gl_hospitality(row: ACRow, context: object) -> object:
    """ Hospitality => column AE = 52204, 52206, 52207, 54211 """
    return _make_pattern("00FFCC50") if row["AE"] in ['52204', '52206', '52207', '54211'] else None

def gl_restricted_memberships(row: ACRow, context: object) -> object:
    """ Restricted Memberships => column AE = 58208, 52336, 52337 """
    return _make_pattern("00FFCC49") if row["AE"] in ['58208', '52336', '52337'] else None

def gl_restricted_travel_status(row: ACRow, context: object) -> object:
    """ Restricted Travel Status Related Expenses => column AE = 54260, 54206, 54208 """ 
    return _make_pattern("00FFCC40") if row["AE"] in ["54260", "54206", "54208"] else None

def gl_restricted_motor_vehicles(row: ACRow, context: object) -> object:
    """ Restricted Motor Vehicles Expenses => column AE = 53332, 53334, 54285, 54287 54000, 54002, 54004  """ 
    return _make_pattern("00FFCC39") if row["AE"] in ["53332", "53334", "54285", "54287" "54000", "54002", "54004"] else None

def gl_restricted_home_equipment_purchases(row: ACRow, context: ACContext) -> object:
    """ Restricted Home Equipment Purchases => AE = 54260 and E >= $800 and CH not in FPS-CHs list """
    return _make_pattern("00EF8B47") if row["AE"] == "54260" and row["E"] >= 800.00 and row["A"] not in context.fps_cardholders else None

###### rules - end ######

@dataclass()
class Criteria:
    name: str
    rule: Callable

criterias = [
    Criteria(name="Restricted IS/OGD transactions", rule=mcc_9399), 
    Criteria(name="Restricted Travel Status Related Expenses (MCC 7011)", rule=mcc_7011),
    Criteria(name="Restricted Travel Status Related Expenses (MCC 4111)", rule=mcc_4111),
    Criteria(name="Restricted Travel Status Related Expenses (MCC 7513)", rule=mcc_7513),
    Criteria(name="Event/Hospitality", rule=mcc_7932),
    Criteria(name="TA Limit exceeded", rule=ta_limit_exceeded),
    Criteria(name="Monitoring of Convenience Cheques", rule=empty_cheques),
    Criteria(name="Financial Coding Error - Designated Science Conferences", rule=gl_financial_coding_error),
    Criteria(name="Conference fees", rule=gl_conference_fees),
    Criteria(name="Hospitality", rule=gl_hospitality),
    Criteria(name="Restricted Memberships", rule=gl_restricted_memberships),
    Criteria(name="Restricted Home Equipment Purchases", rule=gl_restricted_home_equipment_purchases),
    Criteria(name="Restricted Travel Status Related Expenses (GL)", rule=gl_restricted_travel_status),
    Criteria(name="Restricted Motor Vehicles Expenses", rule=gl_restricted_motor_vehicles)
]

def apply_criterias(row, context) -> Criteria:
    """ Applies all the available criterias """
    for criteria in criterias:
        if criteria.rule(row, context):
            return criteria
    return None
