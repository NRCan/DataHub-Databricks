""" CA Processor Main """

import sys
import os
from typing import Callable
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from ac_processor_rules import apply_criterias
from ac_processor_utils import ACRow, ACContext, get_valid_splits, update_splits, get_valid_splits, get_selected_rows

def _get_output_name(file_name: str):
    split = os.path.splitext(file_name)
    return f"{split[0]}_results{split[1]}"

def _highlight_row(row, color):
    for cell in row:
        cell.fill = color

def _load_fps_chs(sheet) -> set[str]:
    values = set()
    for row in sheet.iter_rows(min_row = 2, max_col = 2, max_row = sheet.max_row):
        if row[1].value:
            values.add(row[1].value.strip().upper())
    return values

def _load_rcm_chs(sheet) -> set[str]:
    values = set()
    for row in sheet.iter_rows(min_row = 2, max_col = 3, max_row = sheet.max_row):
        if row[2].value:
            values.add(row[2].value.strip().upper())
    return values

def _load_filter_colors(sheet) -> dict:
    colors = {}
    for row in sheet.iter_rows(min_row = 5, max_col = 3, max_row = sheet.max_row):
        cell = row[2]
        if cell.value:
            color = f"{cell.fill.start_color.rgb}" 
            colors[cell.value.strip()] = PatternFill(start_color=color, end_color=color, fill_type="solid")
    return colors

def _run_processor(argvs):
    if len(argvs) != 2:
        _, tail = os.path.split(argvs[0])
        print(f'Usage:\n\t{tail} <excel_file_name>')
        return 0

    input_file = argvs[1]

    workbook = load_workbook(input_file, read_only=False, data_only=True)
    fps_chs = _load_fps_chs(workbook["FPS CHs"])
    rcm_chs = _load_rcm_chs(workbook["RCM - CH Listing "]) #notice there is a space at the end in the sheet name

    context = ACContext(fps_cardholders=fps_chs, rcm_cardholders=rcm_chs)
    colors = _load_filter_colors(workbook["Filters"])

    master_sheet = workbook.worksheets[0]
    last_column = master_sheet.max_column + 1

    master_sheet.cell(row=1, column=last_column).value = "CRITERIA"

    cardholders = {}

    def _iterate_main_sheet(action: Callable):
        """ iterates the valid part of the main sheet """
        
        row_index = 2
        for excel_row in master_sheet.iter_rows(min_row = row_index, max_col = last_column, max_row = master_sheet.max_row):
            action(excel_row, row_index)
            row_index += 1

    def _process_row(excel_row, index):
        """ highlights a row acording to the given criterias """

        row = ACRow(row=excel_row, index=index)
        cardholder = row["A"]
        if cardholder is not None:
            criteria = apply_criterias(row, context)
            if criteria is not None:
                _highlight_row(excel_row, colors[criteria.name])
                excel_row[last_column - 1].value = criteria.name

            # update splits
            if cardholder not in fps_chs:
                update_splits(cardholders, row)

    # process all rows and potential calculate splits
    _iterate_main_sheet(_process_row)

    splits = get_valid_splits(cardholders, 5000.0)
    split_rows = get_selected_rows(splits)
    potential_splits = "Potential Splits"
    split_pattern = colors[potential_splits]

    def _highlight_splits(excel_row, index):
        if index in split_rows:
            _highlight_row(excel_row, split_pattern)
            excel_row[last_column - 1].value = potential_splits

    # highlight splits
    _iterate_main_sheet(_highlight_splits)
    
    output_file = _get_output_name(input_file)
    workbook.save(output_file)

    print(f"Output file: {output_file}\n")

    return 0

def main():
    #sys.exit(run_processor(sys.argv))
    sys.exit(_run_processor([sys.argv[0], "./data/AC_sample.xlsx"]))

if __name__ == '__main__':
    main()
